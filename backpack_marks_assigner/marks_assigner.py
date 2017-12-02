"""
Script to assign marks from a xlsx to backpack xls.
email column is A, marks column is S.
"""
import openpyxl
import os
import xlrd
import xlwt
from xlutils.copy import copy
import sys

EMAIL_COLUMN = 'A'
MARKS_COLUMN = 'S'
STARTING_MARKS_ROW = 3
LAST_ROW_NO_IN_BP_XLS = 110

if(len(sys.argv) < 3):
    print('Usage:python3 your_marks.xlsx backpack_marks.xls')
    exit()

backpack_row_dict = dict()

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

marks_dict = dict()

# 3 is number from where email list starts, A is the columns no. of emails
for i in range(STARTING_MARKS_ROW,sheet.max_row + 1):
    if(sheet[EMAIL_COLUMN + str(i)].value is not None):
        if(not("GRP" in sheet[EMAIL_COLUMN + str(i)].value and "GRP 9" not in sheet[EMAIL_COLUMN + str(i)].value)):
            marks_dict.setdefault(sheet[EMAIL_COLUMN + str(i)].value,0.0)
            if(sheet[MARKS_COLUMN + str(i)].value is not None):
                marks_dict[sheet[EMAIL_COLUMN + str(i)].value] = sheet[MARKS_COLUMN + str(i)].value

            # for finding integer part of email
            # email_no.append(re.findall(r'\d+', sheet['A' + str(i)].value)[0])

wb.close()
print(marks_dict)

rworkbook = xlrd.open_workbook(sys.argv[2])
rworksheet = rworkbook.sheet_by_index(0)
print(rworksheet.cell(0,0).value)

for email,marks in marks_dict.items():
    for i in range(1,LAST_ROW_NO_IN_BP_XLS):
        if(rworksheet.cell(i,2).value == email):
            backpack_row_dict[email] = i

wworkbook = copy(rworkbook)

# releasing resources for read workbook
# rworkbook.release_resources()
# del rworkbook

wworksheet = wworkbook.get_sheet(0)

print(backpack_row_dict)

for email, marks in marks_dict.items():
    try:
        wworksheet.write(backpack_row_dict[email],3,marks)
    except:
        continue

wworkbook.save('assigned_marks.xls')
